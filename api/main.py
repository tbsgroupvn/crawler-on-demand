from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, HttpUrl
from typing import List, Optional, Dict, Any
import uuid
import json
import os
from datetime import datetime
import redis
from celery import Celery
import databases
import sqlalchemy
from sqlalchemy import create_engine, MetaData, Table, Column, String, DateTime, Text
import re
from collections import Counter
from datetime import datetime, timedelta
import asyncio

# Database setup  
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./app.db")
database = databases.Database(DATABASE_URL)
metadata = MetaData()

# Tasks table
tasks_table = Table(
    "tasks",
    metadata,
    Column("id", String, primary_key=True),
    Column("url", String, nullable=False),
    Column("status", String, nullable=False, default="pending"),
    Column("result", Text, nullable=True),
    Column("error", Text, nullable=True),
    Column("created_at", DateTime, default=datetime.utcnow),
    Column("completed_at", DateTime, nullable=True),
)

engine = create_engine(DATABASE_URL)
metadata.create_all(engine)

# Redis setup
redis_client = redis.from_url(os.getenv("REDIS_URL", "redis://localhost:6379/0"))

# Celery setup
celery_app = Celery(
    "worker",
    broker=os.getenv("REDIS_URL", "redis://localhost:6379/0"),
    backend=os.getenv("REDIS_URL", "redis://localhost:6379/0")
)

# FastAPI app
app = FastAPI(
    title="Crawler On Demand API",
    description="A simple web crawler API",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models
class CrawlRequest(BaseModel):
    url: HttpUrl
    depth: Optional[int] = 1
    max_pages: Optional[int] = 10

class TaskResponse(BaseModel):
    id: str
    url: str
    status: str
    result: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    created_at: datetime
    completed_at: Optional[datetime] = None

# Database connection events
@app.on_event("startup")
async def startup():
    await database.connect()

@app.on_event("shutdown")
async def shutdown():
    await database.disconnect()

# API endpoints
@app.get("/")
async def root():
    return {
        "message": "Crawler On Demand API",
        "status": "running", 
        "version": "1.0.0"
    }

@app.post("/crawl")
async def create_crawl_task(request: CrawlRequest):
    task_id = str(uuid.uuid4())
    
    # Save task to database
    query = tasks_table.insert().values(
        id=task_id,
        url=str(request.url),
        status="pending",
        created_at=datetime.utcnow()
    )
    await database.execute(query)
    
    # Send task to Celery worker
    celery_app.send_task(
        "worker.crawl_url",
        args=[task_id, str(request.url), request.depth, request.max_pages]
    )
    
    return {
        "task_id": task_id,
        "status": "pending",
        "message": "Crawl task created successfully"
    }

@app.get("/tasks/{task_id}")
async def get_task_status(task_id: str):
    query = tasks_table.select().where(tasks_table.c.id == task_id)
    task = await database.fetch_one(query)
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    result = None
    if task.result:
        try:
            result = json.loads(task.result)
        except:
            result = {"raw": task.result}
    
    return {
        "id": task.id,
        "url": task.url,
        "status": task.status,
        "result": result,
        "error": task.error,
        "created_at": task.created_at,
        "completed_at": task.completed_at
    }

@app.get("/tasks")
async def list_tasks(limit: int = 20, offset: int = 0):
    # Get total count
    count_query = "SELECT COUNT(*) FROM tasks"
    total = await database.fetch_val(count_query)
    
    # Get tasks
    query = tasks_table.select().order_by(tasks_table.c.created_at.desc()).limit(limit).offset(offset)
    tasks = await database.fetch_all(query)
    
    task_responses = []
    for task in tasks:
        result = None
        if task.result:
            try:
                result = json.loads(task.result)
            except:
                result = {"raw": task.result}
        
        task_responses.append({
            "id": task.id,
            "url": task.url,
            "status": task.status,
            "result": result,
            "error": task.error,
            "created_at": task.created_at,
            "completed_at": task.completed_at
        })
    
    return {"tasks": task_responses, "total": total}

@app.get("/health")
async def health_check():
    try:
        await database.fetch_val("SELECT 1")
        db_status = "healthy"
    except:
        db_status = "unhealthy"
    
    try:
        redis_client.ping()
        redis_status = "healthy"
    except:
        redis_status = "unhealthy"
    
    return {
        "status": "running",
        "database": db_status,
        "redis": redis_status,
        "celery": "configured"
    }

# New features added!

@app.get("/export/csv/{task_id}")
async def export_task_csv(task_id: str):
    from fastapi.responses import StreamingResponse
    import csv
    import io
    
    query = tasks_table.select().where(tasks_table.c.id == task_id)
    task = await database.fetch_one(query)
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if not task.result:
        raise HTTPException(status_code=400, detail="Task has no results to export")
    
    try:
        result = json.loads(task.result)
    except:
        raise HTTPException(status_code=400, detail="Invalid task result format")
    
    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['URL', 'Title', 'Description', 'Headings', 'Paragraph_Count', 'Link_Count', 'Image_Count'])
    
    # Data rows
    for page in result.get('pages', []):
        writer.writerow([
            page.get('url', ''),
            page.get('title', ''),
            page.get('description', '')[:100] + '...' if len(page.get('description', '')) > 100 else page.get('description', ''),
            '; '.join(page.get('headings', []))[:200],
            len(page.get('paragraphs', [])),
            len(page.get('links', [])),
            len(page.get('images', []))
        ])
    
    output.seek(0)
    
    def iter_csv():
        yield output.getvalue()
    
    return StreamingResponse(
        iter_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=crawl_results_{task_id}.csv"}
    )

@app.get("/export/excel/{task_id}")
async def export_task_excel(task_id: str):
    from fastapi.responses import StreamingResponse
    import io
    import json
    from datetime import datetime
    
    query = tasks_table.select().where(tasks_table.c.id == task_id)
    task = await database.fetch_one(query)
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if not task.result:
        raise HTTPException(status_code=400, detail="Task has no results to export")
    
    try:
        result = json.loads(task.result)
    except:
        raise HTTPException(status_code=400, detail="Invalid task result format")
    
    # Create Excel content using openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Crawl Results"
        
        # Headers with styling
        headers = ['URL', 'Title', 'Description', 'Headings', 'Paragraphs', 'Links', 'Images', 'Content Size', 'Crawled At']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, page in enumerate(result.get('pages', []), 2):
            ws.cell(row=row, column=1, value=page.get('url', ''))
            ws.cell(row=row, column=2, value=page.get('title', ''))
            ws.cell(row=row, column=3, value=page.get('description', '')[:200] + '...' if len(page.get('description', '')) > 200 else page.get('description', ''))
            ws.cell(row=row, column=4, value='; '.join(page.get('headings', []))[:300])
            ws.cell(row=row, column=5, value=len(page.get('paragraphs', [])))
            ws.cell(row=row, column=6, value=len(page.get('links', [])))
            ws.cell(row=row, column=7, value=len(page.get('images', [])))
            ws.cell(row=row, column=8, value=page.get('content_length', 0))
            ws.cell(row=row, column=9, value=page.get('scraped_at', ''))
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        def iter_excel():
            yield excel_buffer.getvalue()
        
        return StreamingResponse(
            iter_excel(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=crawl_results_{task_id}.xlsx"}
        )
        
    except ImportError:
        # Fallback to CSV if openpyxl not available
        raise HTTPException(status_code=501, detail="Excel export requires openpyxl. Use CSV export instead.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")

@app.get("/tasks/filter/{status}")
async def filter_tasks_by_status(status: str, limit: int = 20):
    """Filter tasks by status: pending, running, completed, failed"""
    valid_statuses = ['pending', 'running', 'completed', 'failed']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Use: {', '.join(valid_statuses)}")
    
    query = tasks_table.select().where(tasks_table.c.status == status).order_by(tasks_table.c.created_at.desc()).limit(limit)
    tasks = await database.fetch_all(query)
    
    task_responses = []
    for task in tasks:
        result = None
        if task.result:
            try:
                result = json.loads(task.result)
                # Add summary stats
                if 'pages' in result:
                    result['summary'] = {
                        'total_pages': len(result['pages']),
                        'total_links': sum(len(page.get('links', [])) for page in result['pages']),
                        'total_images': sum(len(page.get('images', [])) for page in result['pages'])
                    }
            except:
                result = {"raw": task.result}
        
        task_responses.append({
            "id": task.id,
            "url": task.url,
            "status": task.status,
            "result": result,
            "error": task.error,
            "created_at": task.created_at,
            "completed_at": task.completed_at
        })
    
    return {"tasks": task_responses, "status_filter": status, "count": len(task_responses)}

@app.delete("/tasks/{task_id}")
async def delete_task(task_id: str):
    """Delete a specific task"""
    query = tasks_table.select().where(tasks_table.c.id == task_id)
    task = await database.fetch_one(query)
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    delete_query = tasks_table.delete().where(tasks_table.c.id == task_id)
    await database.execute(delete_query)
    
    return {"message": f"Task {task_id} deleted successfully"}

@app.post("/batch/crawl")
async def batch_crawl(urls: List[str], depth: Optional[int] = 1, max_pages: Optional[int] = 5):
    """Create multiple crawl tasks at once"""
    if len(urls) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 URLs per batch")
    
    tasks = []
    for url in urls:
        task_id = str(uuid.uuid4())
        
        # Save task to database
        query = tasks_table.insert().values(
            id=task_id,
            url=url,
            status="pending",
            created_at=datetime.utcnow()
        )
        await database.execute(query)
        
        # Send task to Celery worker
        celery_app.send_task(
            "worker.crawl_url",
            args=[task_id, url, depth, max_pages]
        )
        
        tasks.append({
            "task_id": task_id,
            "url": url,
            "status": "pending"
        })
    
    return {
        "message": f"Created {len(tasks)} crawl tasks",
        "tasks": tasks
    }

@app.get("/stats")
async def get_system_stats():
    """Get system statistics"""
    # Count tasks by status
    stats = {}
    for status in ['pending', 'running', 'completed', 'failed']:
        count_query = f"SELECT COUNT(*) FROM tasks WHERE status = '{status}'"
        count = await database.fetch_val(count_query)
        stats[f"{status}_tasks"] = count
    
    # Total tasks
    total_query = "SELECT COUNT(*) FROM tasks"
    stats['total_tasks'] = await database.fetch_val(total_query)
    
    # Recent activity (last 24 hours)
    recent_query = "SELECT COUNT(*) FROM tasks WHERE created_at > datetime('now', '-1 day')"
    try:
        stats['tasks_last_24h'] = await database.fetch_val(recent_query)
    except:
        stats['tasks_last_24h'] = 0
    
    return stats

@app.get("/analytics/{task_id}")
async def get_task_analytics(task_id: str):
    """Phân tích nội dung chi tiết của task"""
    try:
        task = database.get_task(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        
        if not task.get('data'):
            return {"error": "No data to analyze"}
        
        analytics = analyze_content(task['data'])
        return {
            "task_id": task_id,
            "status": task['status'],
            "analytics": analytics,
            "generated_at": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/compare")
async def compare_websites(urls: str):
    """So sánh nội dung giữa nhiều websites"""
    try:
        url_list = [url.strip() for url in urls.split(',')]
        if len(url_list) < 2:
            raise HTTPException(status_code=400, detail="Need at least 2 URLs to compare")
        
        comparison = {}
        for url in url_list:
            # Tìm task mới nhất cho URL này
            tasks = database.get_all_tasks()
            latest_task = None
            for task in tasks:
                if task.get('url') == url and task.get('status') == 'completed':
                    if not latest_task or task.get('created_at', '') > latest_task.get('created_at', ''):
                        latest_task = task
            
            if latest_task and latest_task.get('data'):
                comparison[url] = analyze_content(latest_task['data'])
            else:
                comparison[url] = {"error": "No data found for this URL"}
        
        return {
            "comparison": comparison,
            "summary": generate_comparison_summary(comparison)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/search")
async def search_content(request: dict):
    """Tìm kiếm trong dữ liệu đã crawl"""
    try:
        keyword = request.get('keyword', '').lower()
        search_type = request.get('type', 'all')  # all, title, content, links
        
        if not keyword:
            raise HTTPException(status_code=400, detail="Keyword is required")
        
        tasks = database.get_all_tasks()
        results = []
        
        for task in tasks:
            if task.get('status') != 'completed' or not task.get('data'):
                continue
                
            matches = search_in_task(task, keyword, search_type)
            if matches:
                results.append({
                    "task_id": task['id'],
                    "url": task.get('url'),
                    "crawled_at": task.get('created_at'),
                    "matches": matches
                })
        
        return {
            "keyword": keyword,
            "search_type": search_type,
            "total_results": len(results),
            "results": results[:50]  # Limit to 50 results
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/export/json/{task_id}")
async def export_json(task_id: str):
    """Xuất dữ liệu dạng JSON"""
    try:
        task = database.get_task(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        
        export_data = {
            "task_info": {
                "id": task_id,
                "url": task.get('url'),
                "status": task.get('status'),
                "created_at": task.get('created_at'),
                "completed_at": task.get('completed_at')
            },
            "crawl_data": task.get('data', []),
            "analytics": analyze_content(task.get('data', [])) if task.get('data') else None,
            "export_info": {
                "exported_at": datetime.now().isoformat(),
                "format": "json",
                "version": "1.0"
            }
        }
        
        return Response(
            content=json.dumps(export_data, indent=2, ensure_ascii=False),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=crawl_data_{task_id}.json"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard/advanced")
async def get_advanced_dashboard():
    """Dashboard nâng cao với thống kê chi tiết"""
    try:
        tasks = database.get_all_tasks()
        
        # Thống kê theo thời gian
        time_stats = {}
        domain_stats = {}
        status_stats = {"pending": 0, "running": 0, "completed": 0, "failed": 0}
        content_stats = {"total_pages": 0, "total_content_size": 0, "avg_content_size": 0}
        
        for task in tasks:
            # Thống kê trạng thái
            status = task.get('status', 'unknown')
            status_stats[status] = status_stats.get(status, 0) + 1
            
            # Thống kê theo domain
            url = task.get('url', '')
            if url:
                try:
                    from urllib.parse import urlparse
                    domain = urlparse(url).netloc
                    domain_stats[domain] = domain_stats.get(domain, 0) + 1
                except:
                    pass
            
            # Thống kê nội dung
            data = task.get('data', [])
            if data:
                content_stats["total_pages"] += len(data)
                for page in data:
                    size = page.get('content_size', 0)
                    if isinstance(size, (int, float)):
                        content_stats["total_content_size"] += size
            
            # Thống kê theo thời gian
            created_at = task.get('created_at', '')
            if created_at:
                try:
                    date_key = created_at[:10]  # YYYY-MM-DD
                    time_stats[date_key] = time_stats.get(date_key, 0) + 1
                except:
                    pass
        
        # Tính trung bình
        if content_stats["total_pages"] > 0:
            content_stats["avg_content_size"] = content_stats["total_content_size"] / content_stats["total_pages"]
        
        # Top domains
        top_domains = sorted(domain_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        
        return {
            "summary": {
                "total_tasks": len(tasks),
                "status_breakdown": status_stats,
                "content_stats": content_stats
            },
            "time_series": dict(sorted(time_stats.items())),
            "top_domains": top_domains,
            "generated_at": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Helper functions
def analyze_content(data: list) -> Dict[str, Any]:
    """Phân tích nội dung crawl"""
    if not data:
        return {}
    
    analysis = {
        "total_pages": len(data),
        "total_content_size": 0,
        "avg_content_size": 0,
        "word_count": {"total": 0, "avg_per_page": 0},
        "top_keywords": [],
        "links_analysis": {"total_links": 0, "internal_links": 0, "external_links": 0},
        "content_types": {"with_images": 0, "with_links": 0, "with_headings": 0}
    }
    
    all_text = ""
    total_links = 0
    
    for page in data:
        # Content size
        size = page.get('content_size', 0)
        if isinstance(size, (int, float)):
            analysis["total_content_size"] += size
        
        # Text analysis
        title = page.get('title', '')
        description = page.get('description', '')
        paragraphs = ' '.join(page.get('paragraphs', []))
        headings = ' '.join(page.get('headings', []))
        
        page_text = f"{title} {description} {paragraphs} {headings}"
        all_text += f" {page_text}"
        
        # Links analysis
        links = page.get('links', [])
        total_links += len(links)
        
        # Content types
        if page.get('images'):
            analysis["content_types"]["with_images"] += 1
        if links:
            analysis["content_types"]["with_links"] += 1
        if page.get('headings'):
            analysis["content_types"]["with_headings"] += 1
    
    # Calculate averages
    if analysis["total_pages"] > 0:
        analysis["avg_content_size"] = analysis["total_content_size"] / analysis["total_pages"]
    
    # Word analysis
    words = re.findall(r'\b\w+\b', all_text.lower())
    analysis["word_count"]["total"] = len(words)
    if analysis["total_pages"] > 0:
        analysis["word_count"]["avg_per_page"] = len(words) / analysis["total_pages"]
    
    # Top keywords (excluding common words)
    stop_words = {'the', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'is', 'are', 'was', 'were', 'be', 'been', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must', 'can', 'this', 'that', 'these', 'those', 'a', 'an'}
    filtered_words = [word for word in words if len(word) > 3 and word not in stop_words]
    word_freq = Counter(filtered_words)
    analysis["top_keywords"] = word_freq.most_common(20)
    
    # Links analysis
    analysis["links_analysis"]["total_links"] = total_links
    
    return analysis

def generate_comparison_summary(comparison: Dict) -> Dict:
    """Tạo tóm tắt so sánh"""
    urls = list(comparison.keys())
    summary = {
        "compared_sites": len(urls),
        "metrics": {}
    }
    
    for metric in ["total_pages", "total_content_size", "word_count"]:
        values = []
        for url in urls:
            data = comparison[url]
            if isinstance(data, dict) and metric in data:
                if metric == "word_count":
                    values.append(data[metric].get("total", 0))
                else:
                    values.append(data[metric])
        
        if values:
            summary["metrics"][metric] = {
                "highest": max(values),
                "lowest": min(values),
                "average": sum(values) / len(values)
            }
    
    return summary

def search_in_task(task: dict, keyword: str, search_type: str) -> list:
    """Tìm kiếm trong task"""
    matches = []
    data = task.get('data', [])
    
    for page in data:
        page_matches = {"url": page.get('url'), "matches": []}
        
        if search_type in ['all', 'title']:
            title = page.get('title', '').lower()
            if keyword in title:
                page_matches["matches"].append({"type": "title", "content": page.get('title', '')})
        
        if search_type in ['all', 'content']:
            paragraphs = page.get('paragraphs', [])
            for i, para in enumerate(paragraphs):
                if keyword in para.lower():
                    page_matches["matches"].append({"type": "paragraph", "content": para[:200] + "..."})
        
        if search_type in ['all', 'links']:
            links = page.get('links', [])
            for link in links:
                if keyword in link.lower():
                    page_matches["matches"].append({"type": "link", "content": link})
        
        if page_matches["matches"]:
            matches.append(page_matches)
    
    return matches

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 